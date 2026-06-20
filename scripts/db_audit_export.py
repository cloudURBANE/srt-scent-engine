#!/usr/bin/env python3
"""Live wardrobe DB export + dirty-data audit (read-only).

The monthly "is the data clean?" tool. Connects to the **live app DB** (the
Supabase wardrobe the phone reads), pulls every fragrance out of both
``user_fragrances`` and ``global_fragrances``, flattens each into one clean row,
runs a battery of dirty-data checks, and writes:

  * ``db_export_fragrances.xlsx`` -- one Excel sheet you can open and filter.
    Columns are the facts the app cards depend on; the ``issues`` column lists
    every dirty-data flag for that row (rows with issues are tinted).
  * ``db_export_fragrances.csv``  -- same data, CSV (dependency-free fallback).
  * ``db_audit_report.json``      -- machine-readable issue rollup + per-row flags.

It is strictly READ-ONLY -- it never writes to the DB. To actually *repair* the
flagged rows, run ``scripts/sanitize_wardrobe_accords.py --commit`` (junk +
taxonomy normalization) which shares the same normalization helpers.

DSN: reads ``DATABASE_URL`` from the environment, else from ``huge_monorepo/.env``
(the only DB the app reads). No Railway token needed -- Supabase is off-network.

Usage:
    python scripts/db_audit_export.py                 # export + audit, all rows
    python scripts/db_audit_export.py --owner-only     # only the owner's wardrobe
    python scripts/db_audit_export.py --out my_export   # custom basename
"""
from __future__ import annotations

import argparse
import csv
import datetime as _dt
import hashlib
import json
import os
import re
import sys
import unicodedata
from collections import Counter, defaultdict

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

OWNER = "1c6184cb-0342-48d9-a99a-6cd7018c09ad"

# ---------------------------------------------------------------------------
# Canonical taxonomy (shared with the normalization in dirty_data.py)
# ---------------------------------------------------------------------------
try:
    from dirty_data import (  # type: ignore[import-not-found]
        VALID_CONCENTRATIONS, VALID_GENDERS, audit_row, norm_family,
    )
except ImportError:  # imported as scripts.db_audit_export by tests/tooling
    from scripts.dirty_data import (
        VALID_CONCENTRATIONS, VALID_GENDERS, audit_row, norm_family,
    )


def _resolve_dsn() -> str:
    dsn = os.environ.get("DATABASE_URL")
    if dsn:
        return dsn.strip()
    env_path = os.path.join(ROOT, "..", "huge_monorepo", ".env")
    with open(env_path, encoding="utf-8") as f:
        for line in f:
            m = re.match(r"\s*DATABASE_URL\s*=\s*(.*)\s*$", line)
            if m:
                return m.group(1).strip().strip('"').strip("'")
    raise SystemExit("no DATABASE_URL in env or huge_monorepo/.env")


def _connect(dsn):
    try:
        import psycopg2
    except ImportError:
        import psycopg
        conn = psycopg.connect(dsn)
        conn.read_only = True
        conn.autocommit = True
        return conn
    conn = psycopg2.connect(dsn)
    # Defense in depth: this tool must remain DB-read-only even if a future
    # refactor accidentally adds a mutating statement.
    conn.set_session(readonly=True, autocommit=True)
    return conn


def _load(cur, owner_only: bool):
    """Return list of (source_table, row_id, owner, blob)."""
    out = []
    if owner_only:
        cur.execute(
            "SELECT id, user_id, fragrance_data FROM user_fragrances WHERE user_id = %s",
            (OWNER,),
        )
    else:
        cur.execute("SELECT id, user_id, fragrance_data FROM user_fragrances")
    for rid, uid, blob in cur.fetchall():
        if isinstance(blob, str):
            blob = json.loads(blob)
        out.append(("user_fragrances", str(rid), str(uid), blob or {}))
    if not owner_only:
        # global_fragrances keeps name/brand in top-level columns; fold them into
        # the blob so the export rows are readable (profile_data often omits them).
        cur.execute("SELECT id, name, brand, profile_data FROM global_fragrances")
        for rid, gname, gbrand, blob in cur.fetchall():
            if isinstance(blob, str):
                blob = json.loads(blob)
            blob = blob or {}
            blob.setdefault("name", gname)
            blob.setdefault("brand", gbrand)
            out.append(("global_fragrances", str(rid), "", blob))
    return out


def _accords(blob):
    acc = blob.get("accords")
    if isinstance(acc, list):
        return [str(a) for a in acc]
    return []


def _score(v):
    """Metric fields are sometimes {'score': N, ...} dicts; pull the scalar."""
    if isinstance(v, dict):
        return v.get("score")
    return v


def _flatten(source, rid, owner, blob):
    """One clean export row per fragrance."""
    issues = audit_row(blob)
    dm = blob.get("derived_metrics") if isinstance(blob.get("derived_metrics"), dict) else {}
    return {
        "source": source,
        "row_id": rid,
        "owner": "owner" if owner == OWNER else (owner[:8] if owner else "global"),
        "brand": blob.get("brand") or blob.get("house") or "",
        "name": blob.get("name") or "",
        "year": blob.get("year"),
        "concentration": blob.get("concentration"),
        "family": blob.get("family"),
        "gender": blob.get("gender"),
        "n_accords": len(_accords(blob)),
        "accords": ", ".join(_accords(blob)),
        "performance_score": _score(dm.get("performance_score")),
        "value_score": _score(dm.get("value_score")),
        "community_interest_score": _score(dm.get("community_interest_score")),
        "has_pyramid": bool(((dm.get("notes") or {}) if isinstance(dm.get("notes"), dict) else {}).get("has_pyramid")),
        "source_url": blob.get("source_url") or "",
        "image": (blob.get("imageUrl") or blob.get("image_url") or "")[:80],
        "n_issues": len(issues),
        "issues": "; ".join(issues),
    }


def _identity_key(value):
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = text.encode("ascii", "ignore").decode().casefold()
    return re.sub(r"[^a-z0-9]+", " ", text).strip()


def _append_issue(row, issue):
    flags = row["issues"].split("; ") if row["issues"] else []
    if issue not in flags:
        flags.append(issue)
        row["issues"] = "; ".join(flags)
        row["n_issues"] = len(flags)


def _duplicate_scope(row):
    if row["source"] == "user_fragrances":
        return row["source"], row["owner"]
    return row["source"], "global"


def _annotate_cross_row_issues(rows):
    """Report duplicates within the catalog or one owner's wardrobe only."""
    identities = defaultdict(list)
    source_urls = defaultdict(list)
    for row in rows:
        brand = _identity_key(row["brand"])
        name = _identity_key(row["name"])
        if brand and name:
            identities[(*_duplicate_scope(row), brand, name)].append(row)
        source_url = str(row["source_url"] or "").strip().casefold()
        if source_url:
            source_urls[(*_duplicate_scope(row), source_url)].append(row)

    for group in identities.values():
        if len(group) > 1:
            for row in group:
                _append_issue(row, "duplicate_identity")
    for group in source_urls.values():
        if len(group) > 1:
            for row in group:
                _append_issue(row, "duplicate_source_url")


COLUMNS = [
    "source", "row_id", "owner", "brand", "name", "year", "concentration",
    "family", "gender", "n_accords", "accords", "performance_score",
    "value_score", "community_interest_score", "has_pyramid", "source_url",
    "image", "n_issues", "issues",
]


def _write_csv(rows, path):
    os.makedirs(os.path.dirname(os.path.abspath(path)), exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8-sig") as fh:
        w = csv.DictWriter(fh, fieldnames=COLUMNS)
        w.writeheader()
        for r in rows:
            w.writerow(r)


def _write_xlsx(rows, path):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except Exception:
        return False
    os.makedirs(os.path.dirname(os.path.abspath(path)), exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "fragrances"
    header_fill = PatternFill("solid", fgColor="1F3864")
    header_font = Font(bold=True, color="FFFFFF")
    dirty_fill = PatternFill("solid", fgColor="FCE4D6")
    ws.append([c.replace("_", " ") for c in COLUMNS])
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center")
    for r in rows:
        ws.append([
            v if isinstance(v, (str, int, float, bool, type(None))) else json.dumps(v, default=str)
            for v in (r.get(c) for c in COLUMNS)
        ])
        if r["n_issues"]:
            for cell in ws[ws.max_row]:
                cell.fill = dirty_fill
    # widths
    widths = {"brand": 18, "name": 30, "accords": 40, "issues": 50, "image": 30,
              "source": 16, "row_id": 14, "source_url": 30}
    for i, c in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(c, 13)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{ws.max_row}"
    wb.save(path)
    return True


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--owner-only", action="store_true", help="only the owner's wardrobe rows")
    ap.add_argument("--out", default="db_export_fragrances", help="output basename")
    ap.add_argument(
        "--report",
        help="audit JSON path (default: db_audit_report.json for the standard export, "
        "otherwise <out>.audit.json)",
    )
    args = ap.parse_args()

    dsn = _resolve_dsn()
    conn = _connect(dsn)
    cur = conn.cursor()
    try:
        raw = _load(cur, args.owner_only)
        rows = [_flatten(*r) for r in raw]
        _annotate_cross_row_issues(rows)
    finally:
        cur.close()
        conn.close()

    csv_path = f"{args.out}.csv"
    xlsx_path = f"{args.out}.xlsx"
    _write_csv(rows, csv_path)
    xlsx_ok = _write_xlsx(rows, xlsx_path)

    # ---- summary report ----
    dirty = [r for r in rows if r["n_issues"]]
    issue_counts = Counter()
    for r in dirty:
        for flag in r["issues"].split("; "):
            tag = flag.split(":", 1)[0]
            issue_counts[tag] += 1

    issue_rows = {
        tag: sorted(
            f"{r['source']}:{r['row_id']}"
            for r in dirty
            if any(flag.split(":", 1)[0] == tag for flag in r["issues"].split("; "))
        )
        for tag in issue_counts
    }
    snapshot_payload = json.dumps(issue_rows, sort_keys=True, separators=(",", ":"))
    report = {
        "generated_at": _dt.datetime.utcnow().isoformat() + "Z",
        "db_host": re.sub(r"//[^@]*@", "//***@", dsn).split("?")[0],
        "snapshot_id": hashlib.sha256(snapshot_payload.encode("utf-8")).hexdigest()[:16],
        "total_rows": len(rows),
        "rows_with_issues": len(dirty),
        "issue_counts": dict(issue_counts.most_common()),
        "issue_rows": issue_rows,
        "rows": rows,
    }
    report_path = args.report or (
        "db_audit_report.json"
        if args.out == "db_export_fragrances"
        else f"{args.out}.audit.json"
    )
    os.makedirs(os.path.dirname(os.path.abspath(report_path)), exist_ok=True)
    with open(report_path, "w", encoding="utf-8") as fh:
        json.dump(report, fh, indent=2, default=str)

    print(f"DB: {report['db_host']}")
    print(f"Exported {len(rows)} fragrances -> {csv_path}" + (f" + {xlsx_path}" if xlsx_ok else " (xlsx skipped: openpyxl missing)"))
    print(f"\n=== DIRTY-DATA SUMMARY ===")
    print(f"rows with >=1 issue: {len(dirty)} / {len(rows)}")
    for tag, c in issue_counts.most_common():
        print(f"  {c:>4}  {tag}")
    print(f"\nFull per-row flags + export -> {report_path}")
    print("To repair: python scripts/sanitize_wardrobe_accords.py --commit")


if __name__ == "__main__":
    main()
